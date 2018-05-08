# _*_ coding:utf-8 _*_
#用xlrd读取，用openpyxl写入。xlwt耗时大概为openpyxl的42.3%，效率明显较高。
#----------------------------------------------------------------------------
import xlrd
import openpyxl
import datetime
#import matplotlib
#----------------------------------------------------------------------------


def log(func):
    #写日志
    def wrapper(*args,**kw):
        print('[{}]:call{}()'.format(datetime.datetime.now(),func.__name__))
        return func(*args, **kw)
    return wrapper

def matialize(arg):
    pass

@log
def createxlsx(wbname):
    #创建表文件
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet(title='Spring')
    sheet = wb.create_sheet(title='Autumn')
    wb.save(filename=wbname)

@log
def savetoexcel(data,fields,sheetname,wbname):
    # 写入excel文件中 data 数据，date是list数据类型， fields 表头
    #读取表文件，并保存
    wb = openpyxl.load_workbook(filename=wbname)

    sheet = wb.active #获取当前active的Worksheet
    sheet.title = sheetname #定义active的标签名

    # sheet = wb[sheetname] #读取sheetname
    wb.sheetnames
    field = 1 #定义输入坐标

    for field in range(1,len(fields)+1): #写入表头(起点为r1,c1)
        to_insert=sheet.cell(row=1,column=field,value=str(fields[field-1]))
    row1 = 1
    col1 = 0
    lastrow = len(data)+2
    for row1 in range(2,lastrow):#从表头下写入数据(即r2行开始)
        for col1 in range(1,len(data[row1-2])+1):#c1列开始
            to_insert=sheet.cell(row=row1,column=col1,value=str(data[row1-2][col1-1]))

    sheet[f'A{lastrow+1}'] = 'Record by Hui' #绝对座标赋值
    sheet[f'B{lastrow+1}'] = datetime.datetime.now()
    sheet[f'C{lastrow+1}'] = 'To be Continue'
    sheet.append(['Author', "Time", 'Note'])
    wb.save(filename=wbname)
    print('Success to create {} with sheet \"{}\".'.format(wbname,sheetname))

@log
def readexcel(wbname,sheetname):
    #打开表
    wb = xlrd.open_workbook(filename=wbname)
    print ("表单数量:", wb.nsheets)
    print ("表单名称:", wb.sheet_names())
    # 获取第1个表单
    #sh = wb.sheet_by_index(0)
    sh = wb.sheet_by_name(sheet_name=sheetname)
    print (u"表单 %s 共 %d 行 %d 列" % (sh.name, sh.nrows, sh.ncols))
    print ("第二行第三列:", sh.cell_value(1, 2))
    # 遍历所有表单
    for s in wb.sheets():
        for r in range(s.nrows):
            # 输出指定行
            print (s.row(r))
def testtoexcel(data,fields,sheetname,wbname):
    # 写入excel文件中 data 数据，date是list数据类型， fields 表头
    #读取表文件，并保存
    wb = openpyxl.load_workbook(filename=wbname)
    sheet = wb.active #获取当前active的Worksheet
    sheet.title = sheetname #定义active的标签名
    to_insert=sheet.cell(row=1,column=1,value='Ttte1')
    wb.save(filename=wbname)

def main():
    wbname = './Hideyourheart.xlsx' #当前目录下表文件
    sheetname = 'Snow' #标签
    fields = ['Version','Task','Executor','ExpFinTime','State']
    data =  [['LG6022','DtsCountData','Hei','20180504','Processing'],
    ['LG6122','PythonTeaching','Ei','20180509','Design'],
    ['LG6023','ExcelDesign','i','20180512','Complete']]
    stored_file = wbname #处理后保存文件
    createxlsx(wbname)
    savetoexcel(data,fields,sheetname,wbname)
    readexcel(wbname,sheetname)
    # testtoexcel(data,fields,'sheetname',wbname)
if __name__ == '__main__':
    main()
