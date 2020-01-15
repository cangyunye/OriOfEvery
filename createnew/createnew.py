from dataclasses import dataclass
import paramiko
from csv import DictReader
from openpyxl import load_workbook
from pathlib import PurePath
from time import strptime,mktime
from datetime import datetime
import re
from pathlib import PurePath

@dataclass
class Data_S():
	Operation: str = 'U'
	FilePath: str = '修改文件'
	Revision: int = '版本'
	ChangedAuthor: str = '变更人员'
	ChangedDate: str = 'yyyy-mm-dd hh24:mi:ss'  # strptime(ChangedDate, '%Y-%m-%d %H:%M:%S')


@dataclass
class Data_E():
	# 编号，修改文件，开发人员，日期，回归情况
	Number: int = 0
	FilePaths: str = '文件列表'
	FileType: str = '交付件'
	ChangedAuthor: str = '开发人员'
	TestEngineer: str = '测试人员'
	UpStatus: str = '转测状态'
	# 如果B列为日期格式，那么会自动转换为datetime.datetime对象
	ChangedDate: str = 'yyyy-mm-dd hh24:mi:ss'
	Status: str = '回归情况'

def GetFileName(text:str):
    # 过滤掉特殊字符
    tarray = []
    for i in text.split():
        tarray.append(re.sub('["。”]',"",PurePath(i).name))   
    return tarray


def FindOne(efilename,eFileType,eChangedDate):
    f = lambda x: abs(datetime.strptime(x.ChangedDate,'%Y-%m-%d %H:%M:%S').__rsub__(eChangedDate).total_seconds()) < 3600
    if re.search('脚本',eFileType):
        for s in shellOutList:
            if re.search(efilename,s.FilePath) and f(s):
                return True
    elif re.search('BIN包',eFileType):
        for s in binOutList:
            if re.search(efilename,s.FilePath) and f(s):
                return True
    elif re.search('JAR包',eFileType):
        for s in jarOutList:
            if re.search(efilename,s.FilePath) and f(s):
                return True
    elif re.search('配置',eFileType):
        for s in cfgOutList:
            if re.search(efilename,s.FilePath) and f(s):
                return True   
    else:
        return False

binOutList = []
jarOutList = []
cfgOutList = []
shellOutList = []
with open('F:\\DreamToDream\\OnMyWay\\Programming\\Python\\miniproject\\RefreshConfirm\\1.csv', 'r', encoding='utf-8') as csvfile:
    reader = DictReader(csvfile)
    for row in reader:
        shellOutList.append(
            Data_S(row['Operation'], row['FilePath'], row['Revision'], row['ChangedAuthor'], row['ChangedDate']))

excelOutList = []
wb = load_workbook('F:\\DreamToDream\\OnMyWay\\Programming\\Python\\miniproject\\RefreshConfirm\\刷包.xlsx')
sheet = wb.get_sheet_by_name('Sheet')
for row in map(str, range(2, sheet.max_row + 1)):
    if sheet[f'L{row}'].value == '测试回归完成' or sheet[f'A{row}'].value == None:
        continue
    excelOutList.append(
        Data_E(sheet[f'A{row}'].value, sheet[f'F{row}'].value, sheet[f'E{row}'].value, sheet[f'G{row}'].value,
                sheet[f'H{row}'].value, sheet[f'I{row}'].value, sheet[f'B{row}'].value, sheet[f'L{row}'].value))
wb.close()      

rep_commited = open('F:\\DreamToDream\\OnMyWay\\Programming\\Python\\miniproject\\RefreshConfirm\\rep_commited.txt','w+')
rep_commited.write("=============="*5+"刷包转测文件数据整理，已转测部分"+"============"*5+"\n")
rep_ready = open('F:\\DreamToDream\\OnMyWay\\Programming\\Python\\miniproject\\RefreshConfirm\\rep_ready.txt','w+')
rep_ready.write("============="*5+"刷包转测文件数据整理，登记1小时以上部分"+"============"*5+"\n")



for e in excelOutList[1:]:
    # 打印已转测部分
    if e.UpStatus is not  None and re.search('已转测',e.UpStatus) :
        rep_commited.write(f"编号：{e.Number}\n")
        rep_commited.write(f"文件：{e.FilePaths}\n")
        rep_commited.write(f"测试人员：{e.TestEngineer}\n")
        rep_commited.write(f"说明： excel中登记已转测")
        rep_commited.write("================================="*5+"\n")
    # 对文件类型分类，用于对不同地址获取的svn信息进行确认
    else:       
        for efilename in GetFileName(e.FilePaths):
            # 查找excel中未登记已转测，但svn有最近1小时提交记录的文件
            if FindOne(efilename,e.FileType,e.ChangedDate):
                rep_commited.write(f"编号：{e.Number}\n")
                rep_commited.write(f"文件：{e.FilePaths}\n")
                rep_commited.write(f"测试人员：{e.TestEngineer}\n") 
                rep_commited.write(f"说明： excel中未登记已转测，但svn有最近1小时提交记录的文件")
                rep_commited.write("================================="*5+"\n")
            # excel中未登记已转测，但svn有1小时以上提交记录的文件
            elif e.ChangedDate.__rsub__(datetime.now()).seconds >= 3600:
                rep_commited.write(f"编号：{e.Number}\n")
                rep_commited.write(f"文件：{e.FilePaths}\n")
                rep_commited.write(f"测试人员：{e.TestEngineer}\n")
                rep_commited.write(f"说明： excel中未登记已转测，但svn有1小时以上提交记录的文件")
                rep_commited.write("================================="*5+"\n")
            else:
                rep_ready.write(f"编号：{e.Number}\n")
                rep_ready.write(f"文件：{e.FilePaths}\n")
                rep_ready.write(f"测试人员：{e.TestEngineer}\n")
                rep_ready.write(f"说明： 不完整数据，待确认")   
                rep_ready.write("================================="*5+"\n")

rep_commited.close()
rep_ready.close()

"""
def filematcher(tbd):
	S, iE = tbd
	# 将文件列表分成每个文件
	E_chd = mktime(strptime(iE.ChangedDate))  # 每行的处理时间
	for ie in iE.FilePaths.split('\n'):
		# 将每个文件比对是否在shellOut里出现过
		for iS in S:
			S_chd = mktime(strptime(iS.ChangedDate))
			if re.search(PurePath(ie).name, iS.FilePath) and abs(S_chd - E_chd) <= 3600:
				return True
"""
# 发到指定人员邮箱