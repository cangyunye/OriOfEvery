#!/bin/python3
# 导入配置文件
from .config import *
from dataclasses import dataclass
import paramiko
from csv import DictReader
from openpyxl import load_workbook
from pathlib import PurePath
from datetime import datetime
import re

# 读取配置
HOSTNAME = sshinfo['hostname']
PORT = sshinfo['port']
USERNAME = sshinfo['username']
PASSWORD = sshinfo['password']

# 此处要主要路径转换为绝对路径不能使用~
downloadlist = []  # (remotefile,localfile)
remotepath = report['remotepath']
localpath = report['localpath']
# 调用脚本执行命令
cmd_shell = f'export BASH_ENV=~/.bashrc;. ~/.profile && bash ~/RefreshConfirm/GetSvnLatestInfo.sh "{svnlocalpath["mainupgradepath"]}" "{svnpath["mainupgradepath"]}" '
cmd_cpp = f'export BASH_ENV=~/.bashrc;. ~/.profile && bash ~/RefreshConfirm/GetSvnLatestInfo.sh "{svnlocalpath["bin"]}" "{svnpath["bin"]}" '
cmd_jar = f'export BASH_ENV=~/.bashrc;. ~/.profile && bash ~/RefreshConfirm/GetSvnLatestInfo.sh "{svnlocalpath["jar"]}" "{svnpath["jar"]}" '
cmd_war = f'export BASH_ENV=~/.bashrc;. ~/.profile && bash ~/RefreshConfirm/GetSvnLatestInfo.sh "{svnlocalpath["war"]}" "{svnpath["war"]}" '


# 定义shellOut协议
@dataclass
class Data_S():
	Operation: str = 'U'
	FilePath: str = '修改文件'
	Revision: int = '版本'
	ChangedAuthor: str = '变更人员'
	# strptime(ChangedDate, '%Y-%m-%d %H:%M:%S')
	ChangedDate: str = 'yyyy-mm-dd hh24:mi:ss'


# 定义execleOut协议
@dataclass
class Data_E():
	# 编号，修改文件，开发人员，日期，回归情况
	Number: int = 0
	FilePaths: str = '文件列表'
	FileType: str = '交付件'
	ChangedAuthor: str = '开发人员'
	TestEngineer: str = '测试人员'
	UpStatus: str = '转测状态'
	# 如果B列为日期格式，那么会自动转换为datetime.datetime对象
	ChangedDate: str = 'yyyy-mm-dd hh24:mi:ss'
	Status: str = '回归情况'


def setdata(svnoutfile):
	print(f"追加{svnoutfile}更新文件到Data_S列表")
	svnoutlist = []
	with open(svnoutfile, 'r', encoding='utf-8') as csvfile:
		reader = DictReader(csvfile)
		for row in reader:
			svnoutlist.append(
				Data_S(row['Operation'], row['FilePath'], row['Revision'], row['ChangedAuthor'], row['ChangedDate']))
	return svnoutlist


def GetFileName(text: str):
	# 过滤掉特殊字符
	tarray = []
	for i in text.split():
		tarray.append(re.sub('["。”]', "", PurePath(i).name))
	return tarray


def retrieveAppend(command, ssh):
	print(f"执行后台命令：\ncommand")
	stdin, stdout, stderr = ssh.exec_command(command)
	_file = stdout.read()
	remotefile = _file
	# 过滤为仅文件名不带目录
	localfile = PurePath(localpath, PurePath(_file.decode().strip()).name)
	downloadlist.append((remotefile.decode().strip(), localfile))


def main():
	# 打印配置信息
	# print('svnpath', locals()['svnpath'], 'svnlocalpath', locals()['svnlocalpath'], 'sshinfo', locals()['sshinfo'], 'report',
	#       locals()['report'], sep='\n')
	# 建立与后台的链接对象
	ssh = paramiko.SSHClient()
	ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
	ssh.connect(HOSTNAME, port=PORT, username=USERNAME, password=PASSWORD)
	# 调用后台shell脚本，生成shellOut文件
	retrieveAppend(cmd_shell, ssh)
	# 调用后台shell脚本，生成cppOut文件
	retrieveAppend(cmd_cpp, ssh)
	# 调用后台shell脚本，生成jarOut文件
	retrieveAppend(cmd_jar, ssh)
	# 调用后台shell脚本，生成warOut文件
	retrieveAppend(cmd_war, ssh)
	# 方式1：下载文件
	try:
		trans = ssh.get_transport()
		sftp = paramiko.SFTPClient.from_transport(trans)
		for file in downloadlist:
			print(f"下载文件中...{file[0]}")
			sftp.get(*file)
	except FileNotFoundError as e:
		print("文件不存在，", *file)

	# 方式2：直接返回数据
	# stdin,stdout,stderr=ssh.exec_command('cat '+ remotepath)
	# shellOut=stdout.read()
	# shellOut.decode().split('\n')
	ssh.close()

	# 读取svnout文件，保存为Data_S对象
	shellOutList = setdata(downloadlist[0][1])
	binOutList = setdata(downloadlist[1][1])
	jarOutList = setdata(downloadlist[2][1])
	warOutList = setdata(downloadlist[3][1])

	def FindOne(efilename, eFileType, eChangedDate):
		# 判断文件是否一小时内更新
		def f(x):
			return abs(
				datetime.strptime(x.ChangedDate, '%Y-%m-%d %H:%M:%S').__rsub__(eChangedDate).total_seconds()) < 3600

		if re.search('脚本|配置', eFileType):
			for s in shellOutList:
				if re.search(efilename, s.FilePath) and f(s):
					return True
		elif re.search('BIN包', eFileType):
			for s in binOutList:
				if re.search(efilename, s.FilePath) and f(s):
					return True
		elif re.search('JAR包', eFileType):
			for s in jarOutList:
				if re.search(efilename, s.FilePath) and f(s):
					return True
		elif re.search('WAR包', eFileType):
			for s in warOutList:
				if re.search(efilename, s.FilePath) and f(s):
					return True
		else:
			return False

	# 过滤掉非指定版本的信息
	# filter
	# 读取刷包excel，保存到Data_E对象
	excelOutList = []
	wb = load_workbook(report['rush'])
	# sheet = wb.get_sheet_by_name('Sheet')
	sheet = wb['Sheet']
	for row in map(str, range(2, sheet.max_row + 1)):
		if sheet[f'L{row}'].value == '测试回归完成' or sheet[f'A{row}'].value == None:
			continue
		excelOutList.append(
			Data_E(sheet[f'A{row}'].value, sheet[f'F{row}'].value, sheet[f'E{row}'].value, sheet[f'G{row}'].value,
				   sheet[f'H{row}'].value, sheet[f'I{row}'].value, sheet[f'B{row}'].value, sheet[f'L{row}'].value))
	wb.close()
	# 分析shellOut和excelOut，并输出结果到ReportOut
	rep_commited = open(report['rep1'], 'w+', encoding='utf-8')
	rep_commited.write("==============" * 5 +
					   "刷包转测文件数据整理，已转测部分" + "============" * 5 + "\n")
	rep_ready = open(report['rep2'], 'w+', encoding='utf-8')
	rep_ready.write("=============" * 5 +
					"刷包转测文件数据整理，登记1小时以上部分" + "============" * 5 + "\n")

	for e in excelOutList[1:]:
		# 打印已转测部分
		if e.UpStatus is not None and re.search('已转测', e.UpStatus):
			rep_commited.write(f"编号：{e.Number}\n")
			rep_commited.write(f"文件：{e.FilePaths}\n")
			rep_commited.write(f"测试人员：{e.TestEngineer}\n")
			rep_commited.write(f"说明： excel中登记已转测\n")
			rep_commited.write("=================================" * 5 + "\n")
		# 对文件类型分类，用于对不同地址获取的svn信息进行确认
		else:
			for efilename in GetFileName(e.FilePaths):
				# 查找excel中未登记已转测，但svn有最近1小时提交记录的文件
				if FindOne(efilename, e.FileType, e.ChangedDate):
					rep_commited.write(f"编号：{e.Number}\n")
					rep_commited.write(f"文件：{e.FilePaths}\n")
					rep_commited.write(f"测试人员：{e.TestEngineer}\n")
					rep_commited.write(f"说明： excel中未登记已转测，但svn有最近1小时提交记录的文件\n")
					rep_commited.write(
						"=================================" * 5 + "\n")
					# excel中未登记已转测，但svn有1小时以上提交记录的文件
				elif e.ChangedDate.__rsub__(datetime.now()).seconds >= 3600:
					rep_commited.write(f"编号：{e.Number}\n")
					rep_commited.write(f"文件：{e.FilePaths}\n")
					rep_commited.write(f"测试人员：{e.TestEngineer}\n")
					rep_commited.write(f"说明： excel中未登记已转测，但svn有1小时以上提交记录的文件\n")
					rep_commited.write(
						"=================================" * 5 + "\n")
				else:
					rep_ready.write(f"编号：{e.Number}\n")
					rep_ready.write(f"文件：{e.FilePaths}\n")
					rep_ready.write(f"测试人员：{e.TestEngineer}\n")
					rep_ready.write(f"说明： 不完整数据，待确认\n")
					rep_ready.write(
						"=================================" * 5 + "\n")

	rep_commited.close()
	rep_ready.close()


# 邮件发送
if __name__ == "__main__":
	main()
