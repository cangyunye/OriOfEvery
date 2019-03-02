
# coding: utf-8

# In[1]:


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
wb = Workbook()
#新建一个空白表empty_book.xlsx
dest_filename = 'empty_book.xlsx'
#激活一个worksheet1
ws1 = wb.active
#worksheet1命名
ws1.title = "range names"
#worksheet1扩展40行，每行输入0-599
for row in range(1, 40):
    ws1.append(range(600))
#创建worksheet2,命名Pi
ws2 = wb.create_sheet(title="Pi")
#指定单元格F5的Value=3.14
ws2['F5'] = 3.14
#创建worksheet3,命名Data
ws3 = wb.create_sheet(title="Data")
#在第10行到第20行，27列到54列，填充列索引，并转化为列字母
for row in range(10, 20):
    for col in range(27, 54):
        _ = ws3.cell(column=col, row=row,value="{0}".format(get_column_letter(col)))
print(ws3['AA10'].value)
#隐藏A-D列
ws1.column_dimensions.group('D', 'F', hidden=True)
table = ws1


# In[2]:


# 导入相关模块
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

#字体格式
font0 = Font(name='Calibri',
            size=12,
            bold=True,
            italic=False,
            vertAlign=None,    #Maybe：'baseline', 'superscript', 'subscript'
            underline='none',  #Maybe：'single','double','singleAccounting','doubleAccounting'
            strike=False,
            color='000000')

font1 = Font(name='Microsoft JhengHei',
            size=12,
            bold=True,
            italic=False,
            vertAlign=None,    #Maybe：'baseline', 'superscript', 'subscript'
            underline='none',  #Maybe：'single','double','singleAccounting','doubleAccounting'
            strike=False,
            color='000000')

font2 = Font(name='黑体',
            size=12,
            bold=True,
            italic=False,
            vertAlign=None,    #Maybe：'baseline', 'superscript', 'subscript'
            underline='none',  #Maybe：'single','double','singleAccounting','doubleAccounting'
            strike=False,
            color='FFFFFF')

#单元格填充
fill0 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='C00000',
            end_color='C00000')
fill1 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='FFC000',
            end_color='FFC000')
fill2 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='FFFF00',
            end_color='FFFF00')
fill3 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='92D000',
            end_color='92D000')
fill4 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='00B050',
            end_color='00B050')
fill5 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='00B0F0',
            end_color='00B0F0')
fill6 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='0070C0',
            end_color='0070C0')
fill7 = PatternFill(fill_type='darkGray',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='000000',
            end_color='000000')
fill8 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='7030A0',
            end_color='7030A0')
fill9 = PatternFill(fill_type='lightUp',
#Maybe：'lightUp', 'darkVertical', 'darkGrid', 'solid', 'darkHorizontal', 'darkUp', 'lightVertical', 'lightGray', 'darkTrellis', 'lightDown', 'gray125', 'gray0625', 'mediumGray', 'lightTrellis', 'darkGray', 'darkDown', 'lightHorizontal', 'lightGrid'
            start_color='D2691E',
            end_color='D2691E')

#边框
border0 = Border(left=Side(border_style='thin',color='000000'),
#style Maybe：'mediumDashDotDot', 'dotted', 'thick', 'medium', 'dashDotDot', 'double', 'dashed', 'mediumDashed', 'dashDot', 'mediumDashDot', 'hair', 'slantDashDot', 'thin'
                right=Side(border_style='thin',color='000000'),
                top=Side(border_style='thin',color='000000'),
                bottom=Side(border_style='thin',color='000000'),
                diagonal=Side(border_style=None,color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,color='FF000000'),
                vertical=Side(border_style=None,color='FF000000'),
                horizontal=Side(border_style=None,color='FF000000')
                )
border1 = Border(left=Side(border_style='medium',color='000000'),
#style Maybe：'mediumDashDotDot', 'dotted', 'thick', 'medium', 'dashDotDot', 'double', 'dashed', 'mediumDashed', 'dashDot', 'mediumDashDot', 'hair', 'slantDashDot', 'thin'
                right=Side(border_style='medium',color='000000'),
                top=Side(border_style='medium',color='000000'),
                bottom=Side(border_style='medium',color='000000'),
                diagonal=Side(border_style=None,color='FF000000'),
                diagonal_direction=0,
                outline=Side(border_style=None,color='FF000000'),
                vertical=Side(border_style=None,color='FF000000'),
                horizontal=Side(border_style=None,color='FF000000')
                )

#对齐方式
alignment0 = Alignment(horizontal='general',    #Maybe:'centerContinuous', 'fill', 'right', 'distributed', 'justify', 'general', 'center', 'left'
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)
alignment1 = Alignment(horizontal='right',    #Maybe:'centerContinuous', 'fill', 'right', 'distributed', 'justify', 'general', 'center', 'left'
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=False,
                    shrink_to_fit=False,
                    indent=0)

#表格保护
protection0 = Protection(locked=True,
                         hidden=False)


# In[3]:


# 导入相关模块
from openpyxl.styles import NamedStyle

# 创建格式
style0 = NamedStyle(name = 'style_example0')

#格式赋值
style0.font = font0
style0.alignment = alignment0
style0.border = border0
style0.fill = fill0
style0.Protection = protection0

# 创建格式
style1 = NamedStyle(name = 'style_example1')

#格式赋值
style1.font = font1
style1.alignment = alignment0
style1.border = border0
style1.fill = fill1
style1.Protection = protection0

# 创建格式
style2 = NamedStyle(name = 'style_example2')

#格式赋值
style2.font = font1
style2.alignment = alignment0
style2.border = border0
style2.fill = fill2
style2.Protection = protection0

# 创建格式
style3 = NamedStyle(name = 'style_example3')

#格式赋值
style3.font = font1
style3.alignment = alignment0
style3.border = border0
style3.fill = fill3
style3.Protection = protection0

# 创建格式
style4 = NamedStyle(name = 'style_example4')

#格式赋值
style4.font = font1
style4.alignment = alignment0
style4.border = border0
style4.fill = fill4
style4.Protection = protection0

# 创建格式
style5 = NamedStyle(name = 'style_example5')

#格式赋值
style5.font = font1
style5.alignment = alignment0
style5.border = border0
style5.fill = fill5
style5.Protection = protection0

# 创建格式
style6 = NamedStyle(name = 'style_example6')

#格式赋值
style6.font = font1
style6.alignment = alignment0
style6.border = border0
style6.fill = fill6
style6.Protection = protection0

# 创建格式
style7 = NamedStyle(name = 'style_example7')

#格式赋值
style7.font = font2
style7.alignment = alignment0
style7.border = border0
style7.fill = fill7
style7.Protection = protection0

# 创建格式
style8 = NamedStyle(name = 'style_example8')

#格式赋值
style8.font = font2
style8.alignment = alignment0
style8.border = border0
style8.fill = fill8
style8.Protection = protection0

# 创建格式
style9 = NamedStyle(name = 'style_example9')

#格式赋值
style9.font = font2
style9.alignment = alignment0
style9.border = border0
style9.fill = fill9
style9.Protection = protection0


#格式调用
#单属性调用
# table['A1'].font = font0
# table['A1'].alignment = alignment0
# table.cell(row = 1,column = 1).border = border0

#按名称调用
# table['A1'].style = style0
# table['A1'].style = 'style_example'
# table.cell(row = 1,column = 1).style = style0
# for sty in [style0,style1,style2,style3,style4,style5,style6,style7,style8,style9]:
for col in range(1,10):
    table.cell(row=2,column=col).style=style0
for col in range(1,10):
    table.cell(row=4,column=col).style=style1
for col in range(1,10):
    table.cell(row=6,column=col).style=style2
for col in range(1,10):
    table.cell(row=8,column=col).style=style3
for col in range(1,10):
    table.cell(row=10,column=col).style=style4
for col in range(1,10):
    table.cell(row=12,column=col).style=style5
for col in range(1,10):
    table.cell(row=14,column=col).style=style6
for col in range(1,10):
    table.cell(row=16,column=col).style=style7
for col in range(1,10):
    table.cell(row=18,column=col).style=style8
for col in range(1,10):
    table.cell(row=20,column=col).style=style9


# In[4]:

#Since you're assigning this NamedStyle to more than one cell, it makes sense to register it to your workbook.
wb.add_named_style('style_example1')
wb.add_named_style('style_example2')
wb.add_named_style('style_example3')
wb.add_named_style('style_example4')
wb.add_named_style('style_example5')
wb.add_named_style('style_example6')
wb.add_named_style('style_example7')
wb.add_named_style('style_example8')
wb.add_named_style('style_example9')

wb.save(filename=dest_filename)

